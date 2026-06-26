from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from datetime import date
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from db import database, students, courses, batches, fee_installments
from routes.auth import get_current_user
from utils.helpers import generate_student_id

router = APIRouter(tags=["Students"])


class StudentCreate(BaseModel):
    name: str
    email: Optional[str] = None
    phone: str
    address: Optional[str] = None
    course_id: int
    batch_id: Optional[int] = None
    total_fee: float
    discount: float = 0
    admission_date: date
    status: str = "active"
    notes: Optional[str] = None
    installments: int = 1  # number of installments


@router.get("/students")
async def list_students(
    status: Optional[str] = None,
    course_id: Optional[int] = None,
    batch_id: Optional[int] = None,
    search: Optional[str] = None,
    current_user=Depends(get_current_user)
):
    query = """
        SELECT s.*, c.name as course_name, b.batch_name,
               COALESCE(SUM(col.amount), 0) as total_paid,
               s.net_fee - COALESCE(SUM(col.amount), 0) as balance
        FROM students s
        LEFT JOIN courses c ON s.course_id = c.id
        LEFT JOIN batches b ON s.batch_id = b.id
        LEFT JOIN collections col ON col.student_id = s.id
        WHERE 1=1
    """
    params = {}
    if status:
        query += " AND s.status = :status"
        params["status"] = status
    if course_id:
        query += " AND s.course_id = :course_id"
        params["course_id"] = course_id
    if batch_id:
        query += " AND s.batch_id = :batch_id"
        params["batch_id"] = batch_id
    if search:
        query += " AND (s.name ILIKE :search OR s.phone ILIKE :search OR s.student_id ILIKE :search)"
        params["search"] = f"%{search}%"
    query += " GROUP BY s.id, c.name, b.batch_name ORDER BY s.id DESC"
    return [dict(r) for r in await database.fetch_all(query, params)]


@router.get("/students/{student_id}")
async def get_student(student_id: int, current_user=Depends(get_current_user)):
    query = """
        SELECT s.*, c.name as course_name, b.batch_name,
               COALESCE(SUM(col.amount), 0) as total_paid,
               s.net_fee - COALESCE(SUM(col.amount), 0) as balance
        FROM students s
        LEFT JOIN courses c ON s.course_id = c.id
        LEFT JOIN batches b ON s.batch_id = b.id
        LEFT JOIN collections col ON col.student_id = s.id
        WHERE s.id = :sid
        GROUP BY s.id, c.name, b.batch_name
    """
    row = await database.fetch_one(query, {"sid": student_id})
    if not row:
        raise HTTPException(status_code=404, detail="Student not found")
    
    installments = await database.fetch_all(
        fee_installments.select().where(fee_installments.c.student_id == student_id)
        .order_by(fee_installments.c.installment_no)
    )
    result = dict(row)
    result["installments"] = [dict(i) for i in installments]
    return result


@router.post("/students")
async def create_student(data: StudentCreate, current_user=Depends(get_current_user)):
    # Auto generate student ID
    count = await database.fetch_one("SELECT COUNT(*) as cnt FROM students")
    sid = generate_student_id((count["cnt"] or 0) + 1)
    net_fee = data.total_fee - data.discount

    student_id = await database.execute(students.insert().values(
        student_id=sid,
        name=data.name,
        email=data.email,
        phone=data.phone,
        address=data.address,
        course_id=data.course_id,
        batch_id=data.batch_id,
        total_fee=data.total_fee,
        discount=data.discount,
        net_fee=net_fee,
        admission_date=data.admission_date,
        status=data.status,
        notes=data.notes,
    ))

    # Create installments
    if data.installments > 0:
        per_installment = round(net_fee / data.installments, 2)
        from datetime import timedelta
        for i in range(1, data.installments + 1):
            due = data.admission_date.replace(day=1)
            from dateutil.relativedelta import relativedelta
            try:
                due = due + relativedelta(months=i - 1)
            except Exception:
                pass
            await database.execute(fee_installments.insert().values(
                student_id=student_id,
                installment_no=i,
                amount_due=per_installment,
                due_date=due,
                status="pending",
            ))

    return {"message": "Student created", "id": student_id, "student_id": sid}


@router.put("/students/{student_id}")
async def update_student(student_id: int, data: dict, current_user=Depends(get_current_user)):
    data.pop("installments", None)
    data.pop("id", None)
    if "net_fee" not in data and "total_fee" in data and "discount" in data:
        data["net_fee"] = data["total_fee"] - data["discount"]
    await database.execute(students.update().where(students.c.id == student_id).values(**data))
    return {"message": "Updated"}


# ── EXCEL IMPORT ───────────────────────────────────────────────────────────────

@router.post("/students/import/excel")
async def import_students_excel(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [str(cell.value).strip().lower() for cell in ws[1]]

    created, errors = 0, []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            row_data = dict(zip(headers, row))
            if not row_data.get("name") or not row_data.get("phone"):
                continue

            course_name = str(row_data.get("course", "")).strip()
            course = await database.fetch_one(
                courses.select().where(courses.c.name.ilike(f"%{course_name}%"))
            )
            if not course:
                errors.append(f"Row {row_idx}: Course '{course_name}' not found")
                continue

            count = await database.fetch_one("SELECT COUNT(*) as cnt FROM students")
            sid = generate_student_id((count["cnt"] or 0) + 1)
            total_fee = float(row_data.get("total_fee", course["total_fee"]))
            discount = float(row_data.get("discount", 0))

            await database.execute(students.insert().values(
                student_id=sid,
                name=str(row_data["name"]).strip(),
                email=str(row_data.get("email", "") or ""),
                phone=str(row_data["phone"]).strip(),
                address=str(row_data.get("address", "") or ""),
                course_id=course["id"],
                total_fee=total_fee,
                discount=discount,
                net_fee=total_fee - discount,
                admission_date=date.today(),
                status="active",
            ))
            created += 1
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    return {"created": created, "errors": errors}


# ── EXCEL EXPORT ───────────────────────────────────────────────────────────────

@router.get("/students/export/excel")
async def export_students_excel(current_user=Depends(get_current_user)):
    rows = await database.fetch_all("""
        SELECT s.student_id, s.name, s.phone, s.email, c.name as course,
               b.batch_name as batch, s.total_fee, s.discount, s.net_fee,
               COALESCE(SUM(col.amount),0) as paid,
               s.net_fee - COALESCE(SUM(col.amount),0) as balance,
               s.admission_date, s.status
        FROM students s
        LEFT JOIN courses c ON s.course_id = c.id
        LEFT JOIN batches b ON s.batch_id = b.id
        LEFT JOIN collections col ON col.student_id = s.id
        GROUP BY s.id, c.name, b.batch_name ORDER BY s.id
    """)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["Student ID", "Name", "Phone", "Email", "Course", "Batch",
               "Total Fee", "Discount", "Net Fee", "Paid", "Balance",
               "Admission Date", "Status"]

    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(list(dict(row).values()))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename=students_export.xlsx"})


# ── EXCEL TEMPLATE ─────────────────────────────────────────────────────────────

@router.get("/students/import/template")
async def download_import_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"

    headers = ["name", "phone", "email", "course", "total_fee", "discount", "address"]
    sample = ["Rahul Kumar", "9876543210", "rahul@email.com",
              "Python Full Stack with AI", "35000", "0", "Kochi, Kerala"]

    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    ws.append(sample)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename=import_template.xlsx"})
