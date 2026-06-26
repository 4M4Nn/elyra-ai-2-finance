from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from datetime import date
import io, openpyxl
from openpyxl.styles import Font, PatternFill
from db import database, expenses, salary_records
from routes.auth import get_current_user

router = APIRouter(tags=["Expenses & Salary"])


class ExpenseCreate(BaseModel):
    category: str
    description: str
    amount: float
    expense_date: date
    payment_mode: str
    vendor: Optional[str] = None
    reference: Optional[str] = None


class SalaryCreate(BaseModel):
    staff_name: str
    designation: Optional[str] = None
    month: int
    year: int
    gross_salary: float
    deductions: float = 0
    paid_date: Optional[date] = None
    payment_mode: Optional[str] = None


# ── EXPENSES ───────────────────────────────────────────────────────────────────

@router.get("/expenses")
async def list_expenses(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    category: Optional[str] = None,
    current_user=Depends(get_current_user)
):
    query = """
        SELECT e.*, u.name as added_by_name
        FROM expenses e LEFT JOIN users u ON e.added_by = u.id
        WHERE 1=1
    """
    params = {}
    if from_date:
        query += " AND e.expense_date >= :from_date"
        params["from_date"] = from_date
    if to_date:
        query += " AND e.expense_date <= :to_date"
        params["to_date"] = to_date
    if category:
        query += " AND e.category = :category"
        params["category"] = category
    query += " ORDER BY e.expense_date DESC, e.id DESC"
    return [dict(r) for r in await database.fetch_all(query, params)]


@router.post("/expenses")
async def add_expense(data: ExpenseCreate, current_user=Depends(get_current_user)):
    eid = await database.execute(expenses.insert().values(
        **data.dict(), added_by=current_user["id"]
    ))
    return {"message": "Expense added", "id": eid}


@router.put("/expenses/{expense_id}")
async def update_expense(expense_id: int, data: ExpenseCreate, current_user=Depends(get_current_user)):
    await database.execute(expenses.update().where(expenses.c.id == expense_id).values(**data.dict()))
    return {"message": "Updated"}


@router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: int, current_user=Depends(get_current_user)):
    await database.execute(expenses.delete().where(expenses.c.id == expense_id))
    return {"message": "Deleted"}


@router.get("/expenses/summary/category")
async def expense_by_category(
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user=Depends(get_current_user)
):
    query = "SELECT category, SUM(amount) as total FROM expenses WHERE 1=1"
    params = {}
    if month:
        query += " AND EXTRACT(MONTH FROM expense_date) = :month"
        params["month"] = month
    if year:
        query += " AND EXTRACT(YEAR FROM expense_date) = :year"
        params["year"] = year
    query += " GROUP BY category ORDER BY total DESC"
    return [dict(r) for r in await database.fetch_all(query, params)]


@router.get("/expenses/export/excel")
async def export_expenses(current_user=Depends(get_current_user)):
    rows = await database.fetch_all("""
        SELECT e.category, e.description, e.amount, e.expense_date,
               e.payment_mode, e.vendor, e.reference, u.name as added_by
        FROM expenses e LEFT JOIN users u ON e.added_by = u.id
        ORDER BY e.expense_date DESC
    """)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    headers = ["Category", "Description", "Amount (₹)", "Date", "Mode", "Vendor", "Reference", "Added By"]
    hfill = PatternFill("solid", fgColor="6366F1")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hfill
        cell.font = Font(bold=True, color="FFFFFF")
    for row in rows:
        ws.append(list(dict(row).values()))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=expenses_export.xlsx"})


# ── SALARY ─────────────────────────────────────────────────────────────────────

@router.get("/salary")
async def list_salary(
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user=Depends(get_current_user)
):
    query = "SELECT s.*, u.name as added_by_name FROM salary_records s LEFT JOIN users u ON s.added_by = u.id WHERE 1=1"
    params = {}
    if month:
        query += " AND s.month = :month"
        params["month"] = month
    if year:
        query += " AND s.year = :year"
        params["year"] = year
    query += " ORDER BY s.year DESC, s.month DESC"
    return [dict(r) for r in await database.fetch_all(query, params)]


@router.post("/salary")
async def add_salary(data: SalaryCreate, current_user=Depends(get_current_user)):
    net = data.gross_salary - data.deductions
    sid = await database.execute(salary_records.insert().values(
        **data.dict(), net_salary=net, added_by=current_user["id"]
    ))
    return {"message": "Salary record added", "id": sid}


@router.put("/salary/{record_id}")
async def update_salary(record_id: int, data: SalaryCreate, current_user=Depends(get_current_user)):
    net = data.gross_salary - data.deductions
    await database.execute(salary_records.update().where(salary_records.c.id == record_id)
                           .values(**data.dict(), net_salary=net))
    return {"message": "Updated"}
