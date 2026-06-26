from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import date, datetime
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from db import database
from routes.auth import get_current_user

router = APIRouter(tags=["Reports & Dashboard"])


@router.get("/dashboard/stats")
async def dashboard_stats(current_user=Depends(get_current_user)):
    today = date.today()
    month = today.month
    year = today.year

    # Today's collection
    today_col = await database.fetch_one(
        "SELECT COALESCE(SUM(amount), 0) as total FROM collections WHERE payment_date = :d",
        {"d": today}
    )

    # Month collection
    month_col = await database.fetch_one(
        "SELECT COALESCE(SUM(amount), 0) as total FROM collections WHERE EXTRACT(MONTH FROM payment_date)=:m AND EXTRACT(YEAR FROM payment_date)=:y",
        {"m": month, "y": year}
    )

    # Month expenses
    month_exp = await database.fetch_one(
        "SELECT COALESCE(SUM(amount), 0) as total FROM expenses WHERE EXTRACT(MONTH FROM expense_date)=:m AND EXTRACT(YEAR FROM expense_date)=:y",
        {"m": month, "y": year}
    )

    # Total outstanding
    outstanding = await database.fetch_one(
        "SELECT COALESCE(SUM(net_fee), 0) - COALESCE((SELECT SUM(amount) FROM collections), 0) as total FROM students WHERE status='active'"
    )

    # Active students
    active_students = await database.fetch_one("SELECT COUNT(*) as cnt FROM students WHERE status='active'")

    # Overdue installments
    overdue = await database.fetch_one(
        "SELECT COUNT(*) as cnt, COALESCE(SUM(amount_due - amount_paid), 0) as total FROM fee_installments WHERE status IN ('pending','partial') AND due_date < :today",
        {"today": today}
    )

    # Monthly trend (last 6 months)
    trend = await database.fetch_all("""
        SELECT TO_CHAR(payment_date, 'Mon YYYY') as month_label,
               EXTRACT(YEAR FROM payment_date) as yr,
               EXTRACT(MONTH FROM payment_date) as mn,
               SUM(amount) as collections
        FROM collections
        WHERE payment_date >= (CURRENT_DATE - INTERVAL '6 months')
        GROUP BY month_label, yr, mn ORDER BY yr, mn
    """)

    # Expense trend
    exp_trend = await database.fetch_all("""
        SELECT TO_CHAR(expense_date, 'Mon YYYY') as month_label,
               EXTRACT(YEAR FROM expense_date) as yr,
               EXTRACT(MONTH FROM expense_date) as mn,
               SUM(amount) as expenses
        FROM expenses
        WHERE expense_date >= (CURRENT_DATE - INTERVAL '6 months')
        GROUP BY month_label, yr, mn ORDER BY yr, mn
    """)

    # Course-wise revenue
    course_revenue = await database.fetch_all("""
        SELECT c.name as course_name, COUNT(s.id) as students,
               SUM(s.net_fee) as total_fees,
               COALESCE(SUM(col.amount), 0) as collected
        FROM courses c
        LEFT JOIN students s ON s.course_id = c.id
        LEFT JOIN collections col ON col.student_id = s.id
        GROUP BY c.name ORDER BY collected DESC
    """)

    # Payment mode breakdown this month
    mode_breakdown = await database.fetch_all("""
        SELECT payment_mode, SUM(amount) as total
        FROM collections
        WHERE EXTRACT(MONTH FROM payment_date)=:m AND EXTRACT(YEAR FROM payment_date)=:y
        GROUP BY payment_mode
    """, {"m": month, "y": year})

    return {
        "today_collection": today_col["total"],
        "month_collection": month_col["total"],
        "month_expenses": month_exp["total"],
        "month_profit": (month_col["total"] or 0) - (month_exp["total"] or 0),
        "total_outstanding": max(outstanding["total"] or 0, 0),
        "active_students": active_students["cnt"],
        "overdue_count": overdue["cnt"],
        "overdue_amount": overdue["total"],
        "monthly_trend": [dict(r) for r in trend],
        "expense_trend": [dict(r) for r in exp_trend],
        "course_revenue": [dict(r) for r in course_revenue],
        "mode_breakdown": [dict(r) for r in mode_breakdown],
    }


@router.get("/reports/pnl")
async def pnl_report(
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user=Depends(get_current_user)
):
    today = date.today()
    m = month or today.month
    y = year or today.year

    collections_total = await database.fetch_one(
        "SELECT COALESCE(SUM(amount), 0) as total FROM collections WHERE EXTRACT(MONTH FROM payment_date)=:m AND EXTRACT(YEAR FROM payment_date)=:y",
        {"m": m, "y": y}
    )
    expenses_by_cat = await database.fetch_all(
        "SELECT category, SUM(amount) as total FROM expenses WHERE EXTRACT(MONTH FROM expense_date)=:m AND EXTRACT(YEAR FROM expense_date)=:y GROUP BY category ORDER BY total DESC",
        {"m": m, "y": y}
    )
    salary_total = await database.fetch_one(
        "SELECT COALESCE(SUM(net_salary), 0) as total FROM salary_records WHERE month=:m AND year=:y",
        {"m": m, "y": y}
    )
    total_expenses = sum(r["total"] for r in expenses_by_cat) + (salary_total["total"] or 0)

    return {
        "month": m, "year": y,
        "total_income": collections_total["total"],
        "expenses_breakdown": [dict(r) for r in expenses_by_cat],
        "salary_total": salary_total["total"],
        "total_expenses": total_expenses,
        "net_profit": (collections_total["total"] or 0) - total_expenses,
    }


@router.get("/reports/outstanding")
async def outstanding_report(current_user=Depends(get_current_user)):
    rows = await database.fetch_all("""
        SELECT s.student_id, s.name, s.phone, c.name as course_name,
               s.net_fee, COALESCE(SUM(col.amount), 0) as paid,
               s.net_fee - COALESCE(SUM(col.amount), 0) as balance
        FROM students s
        LEFT JOIN courses c ON s.course_id = c.id
        LEFT JOIN collections col ON col.student_id = s.id
        WHERE s.status = 'active'
        GROUP BY s.id, c.name
        HAVING s.net_fee - COALESCE(SUM(col.amount), 0) > 0
        ORDER BY balance DESC
    """)
    return [dict(r) for r in rows]


@router.get("/reports/overdue-installments")
async def overdue_installments(current_user=Depends(get_current_user)):
    today = date.today()
    rows = await database.fetch_all("""
        SELECT fi.*, s.name as student_name, s.phone, s.student_id as student_code,
               c.name as course_name
        FROM fee_installments fi
        LEFT JOIN students s ON fi.student_id = s.id
        LEFT JOIN courses c ON s.course_id = c.id
        WHERE fi.status IN ('pending','partial') AND fi.due_date < :today
        ORDER BY fi.due_date ASC
    """, {"today": today})
    return [dict(r) for r in rows]


@router.get("/reports/pnl/export/excel")
async def export_pnl_excel(
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user=Depends(get_current_user)
):
    today = date.today()
    m = month or today.month
    y = year or today.year

    pnl = await pnl_report(month=m, year=y, current_user=current_user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"P&L {m}-{y}"

    hfill = PatternFill("solid", fgColor="6366F1")

    ws.append(["Elyra AI 2 — P&L Report", f"{m}/{y}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["INCOME"])
    ws["A3"].font = Font(bold=True)
    ws.append(["Total Collections", pnl["total_income"]])
    ws.append([])

    ws.append(["EXPENSES"])
    ws["A6"].font = Font(bold=True)
    for row in pnl["expenses_breakdown"]:
        ws.append([row["category"].title(), row["total"]])
    ws.append(["Salary", pnl["salary_total"]])
    ws.append(["Total Expenses", pnl["total_expenses"]])
    ws.append([])
    ws.append(["NET PROFIT / LOSS", pnl["net_profit"]])

    last = ws.max_row
    ws[f"A{last}"].font = Font(bold=True, size=12)
    ws[f"B{last}"].font = Font(bold=True, size=12,
                                color="00AA00" if pnl["net_profit"] >= 0 else "CC0000")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pnl_{m}_{y}.xlsx"})
