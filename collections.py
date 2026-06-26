from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from typing import Optional
from datetime import date, datetime
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from db import database, collections, students, fee_installments, courses, batches
from routes.auth import get_current_user
from utils.helpers import generate_receipt_no
from utils.pdf_receipt import generate_receipt_pdf

router = APIRouter(tags=["Collections"])


class CollectionCreate(BaseModel):
    student_id: int
    installment_id: Optional[int] = None
    amount: float
    payment_mode: str
    payment_date: date
    transaction_ref: Optional[str] = None
    notes: Optional[str] = None


@router.get("/collections")
async def list_collections(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    payment_mode: Optional[str] = None,
    student_id: Optional[int] = None,
    current_user=Depends(get_current_user)
):
    query = """
        SELECT col.*, s.name as student_name, s.student_id as student_code,
               c.name as course_name, u.name as collected_by_name
        FROM collections col
        LEFT JOIN students s ON col.student_id = s.id
        LEFT JOIN courses c ON s.course_id = c.id
        LEFT JOIN users u ON col.collected_by = u.id
        WHERE 1=1
    """
    params = {}
    if from_date:
        query += " AND col.payment_date >= :from_date"
        params["from_date"] = from_date
    if to_date:
        query += " AND col.payment_date <= :to_date"
        params["to_date"] = to_date
    if payment_mode:
        query += " AND col.payment_mode = :payment_mode"
        params["payment_mode"] = payment_mode
    if student_id:
        query += " AND col.student_id = :student_id"
        params["student_id"] = student_id
    query += " ORDER BY col.payment_date DESC, col.id DESC"
    return [dict(r) for r in await database.fetch_all(query, params)]


@router.get("/collections/daily-summary")
async def daily_summary(target_date: Optional[date] = None, current_user=Depends(get_current_user)):
    d = target_date or date.today()
    query = """
        SELECT payment_mode, COUNT(*) as count, SUM(amount) as total
        FROM collections
        WHERE payment_date = :d
        GROUP BY payment_mode
    """
    rows = await database.fetch_all(query, {"d": d})
    total_row = await database.fetch_one(
        "SELECT COUNT(*) as count, SUM(amount) as total FROM collections WHERE payment_date = :d",
        {"d": d}
    )
    return {
        "date": str(d),
        "breakdown": [dict(r) for r in rows],
        "total_collections": total_row["total"] or 0,
        "total_transactions": total_row["count"] or 0,
    }


@router.post("/collections")
async def record_collection(data: CollectionCreate, current_user=Depends(get_current_user)):
    # Validate student
    student = await database.fetch_one(students.select().where(students.c.id == data.student_id))
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    receipt_no = generate_receipt_no()

    col_id = await database.execute(collections.insert().values(
        receipt_no=receipt_no,
        student_id=data.student_id,
        installment_id=data.installment_id,
        amount=data.amount,
        payment_mode=data.payment_mode,
        payment_date=data.payment_date,
        transaction_ref=data.transaction_ref,
        collected_by=current_user["id"],
        notes=data.notes,
    ))

    # Update installment status
    if data.installment_id:
        inst = await database.fetch_one(
            fee_installments.select().where(fee_installments.c.id == data.installment_id)
        )
        if inst:
            new_paid = (inst["amount_paid"] or 0) + data.amount
            new_status = "paid" if new_paid >= inst["amount_due"] else "partial"
            await database.execute(
                fee_installments.update().where(fee_installments.c.id == data.installment_id)
                .values(amount_paid=new_paid, paid_date=data.payment_date, status=new_status)
            )

    return {"message": "Collection recorded", "id": col_id, "receipt_no": receipt_no}


@router.get("/collections/{collection_id}/receipt/pdf")
async def download_receipt(collection_id: int, current_user=Depends(get_current_user)):
    col = await database.fetch_one("""
        SELECT col.*, s.name as student_name, s.student_id as student_code,
               c.name as course_name, b.batch_name,
               s.net_fee - COALESCE((SELECT SUM(amount) FROM collections WHERE student_id = s.id), 0) as balance
        FROM collections col
        LEFT JOIN students s ON col.student_id = s.id
        LEFT JOIN courses c ON s.course_id = c.id
        LEFT JOIN batches b ON s.batch_id = b.id
        WHERE col.id = :cid
    """, {"cid": collection_id})

    if not col:
        raise HTTPException(status_code=404, detail="Collection not found")

    inst_label = "—"
    if col["installment_id"]:
        inst = await database.fetch_one(
            fee_installments.select().where(fee_installments.c.id == col["installment_id"])
        )
        if inst:
            inst_label = f"Installment {inst['installment_no']}"

    pdf_bytes = generate_receipt_pdf({
        "receipt_no": col["receipt_no"],
        "payment_date": str(col["payment_date"]),
        "student_id": col["student_code"],
        "student_name": col["student_name"],
        "course": col["course_name"],
        "batch": col["batch_name"] or "—",
        "amount": col["amount"],
        "payment_mode": col["payment_mode"],
        "transaction_ref": col["transaction_ref"] or "",
        "installment_label": inst_label,
        "description": "Course Fee Payment",
        "outstanding_balance": max(col["balance"] or 0, 0),
    })

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=receipt_{col['receipt_no']}.pdf"}
    )


@router.get("/collections/export/excel")
async def export_collections_excel(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    current_user=Depends(get_current_user)
):
    query = """
        SELECT col.receipt_no, s.student_id as student_code, s.name as student_name,
               c.name as course_name, col.amount, col.payment_mode,
               col.payment_date, col.transaction_ref, u.name as collected_by
        FROM collections col
        LEFT JOIN students s ON col.student_id = s.id
        LEFT JOIN courses c ON s.course_id = c.id
        LEFT JOIN users u ON col.collected_by = u.id
        WHERE 1=1
    """
    params = {}
    if from_date:
        query += " AND col.payment_date >= :from_date"
        params["from_date"] = from_date
    if to_date:
        query += " AND col.payment_date <= :to_date"
        params["to_date"] = to_date
    query += " ORDER BY col.payment_date DESC"

    rows = await database.fetch_all(query, params)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collections"

    headers = ["Receipt No", "Student ID", "Student Name", "Course",
               "Amount (₹)", "Payment Mode", "Date", "Transaction Ref", "Collected By"]

    hfill = PatternFill("solid", fgColor="6366F1")
    hfont = Font(bold=True, color="FFFFFF")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hfill
        cell.font = hfont

    for row in rows:
        ws.append(list(dict(row).values()))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=collections_export.xlsx"})
